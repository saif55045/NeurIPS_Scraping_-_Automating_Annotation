import os
import openpyxl
import google.generativeai as genai

# Set up Google Gemini API
API_KEY = "AIzaSyC_UYJdaED_LSlHJqzSbk_z6mciNv-LSvM"  # Replace with your actual API key
genai.configure(api_key=API_KEY)

# Define categories
CATEGORIES = [
    "Deep Learning",
    "Natural Language Processing",
    "Computer Vision",
    "Reinforcement Learning",
    "Optimization & Theory"
]

# Load Excel file
def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    return wb, sheet

# Ask Gemini to classify the paper
def classify_paper(title, abstract):
    """Ask Gemini to classify the paper"""
    prompt = f"""
    Given the following research paper details:
    Title: {title}
    Abstract: {abstract}
    
    Classify this paper into one of these categories: {', '.join(CATEGORIES)}.
    Just return the category name.
    """
    
    try:
        model = genai.GenerativeModel('gemini-pro')
        response = model.generate_content(prompt)
        category = response.text.strip()
        return category if category in CATEGORIES else "Uncategorized"
    except Exception as e:
        print(f"Error classifying paper: {e}")
        raise e  # Raise exception to handle API limits

# Annotate dataset
def annotate_papers(file_path):
    wb, sheet = load_excel(file_path)
    
    # Add Category column if not present
    if sheet.cell(row=1, column=sheet.max_column).value != "Category":
        sheet.cell(row=1, column=sheet.max_column + 1, value="Category")
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
        title = row[2].value  # Column 3 (Title)
        abstract = row[4].value  # Column 5 (Abstract)
        
        if not title or not abstract:
            category = "Missing Data"
        else:
            try:
                category = classify_paper(title, abstract)
            except Exception as e:
                print("API limit reached or error encountered. Exiting.")
                wb.save(file_path)  # Save progress before exiting
                return
        
        row[sheet.max_column - 1].value = category  # Assign category
        wb.save(file_path)  # Save after each entry
        print(f"Annotated: {title} -> {category}")
    
    print("✅ Annotation complete! File updated.")

# Run annotation
EXCEL_FILE = "NeurIPS_Years/NeurIPS_2024.xlsx"  # Replace with your actual file
annotate_papers(EXCEL_FILE)
